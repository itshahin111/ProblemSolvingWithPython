from openpyxl import load_workbook


def copy_and_delete_data(file_path_1, file_path_2, sheet_name_1, sheet_name_2):
    # Load both workbooks and their sheets
    workbook_1 = load_workbook(file_path_1)
    sheet_1 = workbook_1[sheet_name_1]

    workbook_2 = load_workbook(file_path_2)
    sheet_2 = workbook_2[sheet_name_2]

    # Iterate through each row in the sheet of the first workbook
    max_row = sheet_1.max_row
    row_data = []  # To store the formatted data to write to the second file

    for row in range(1, max_row + 1):
        # Extract values from columns A (company_name), B (phone), C (email), and D (website)
        company_name = sheet_1[f"A{row}"].value or ""
        phone = sheet_1[f"B{row}"].value or ""
        email = sheet_1[f"C{row}"].value or ""
        website = sheet_1[f"D{row}"].value or ""

        # Format the data
        formatted_data = {
            "company_name": company_name,
            "phone": phone,
            "email": email,
            "website": website,
        }

        # Append formatted data to the list to write to the second file
        row_data.append(formatted_data)

        # Delete the row after copying the data
        sheet_1.delete_rows(row)

    # Write the formatted data to the second Excel file
    for index, data in enumerate(row_data, start=1):
        sheet_2[f"A{index}"] = f"company_name => {data['company_name']}"
        sheet_2[f"B{index}"] = f"phone => {data['phone']}"
        sheet_2[f"C{index}"] = f"email => {data['email']}"
        sheet_2[f"D{index}"] = f"website => {data['website']}"

    # Save the updated files
    workbook_1.save(file_path_1)
    workbook_2.save(file_path_2)

    print(
        f"Data has been copied and deleted from the sheet '{sheet_name_1}' to '{sheet_name_2}'."
    )


# Example usage
file_path_1 = "BCD_DataAppendA.xlsx"  # Replace with the path to your first Excel file
file_path_2 = "BaccoAll.xlsx"  # Replace with the path to your second Excel file
sheet_name_1 = "Sheet1"  # Replace with the sheet name in the first file
sheet_name_2 = "Sheet1"  # Replace with the sheet name in the second file

copy_and_delete_data(file_path_1, file_path_2, sheet_name_1, sheet_name_2)

from openpyxl import load_workbook


def append_columns_bcd_to_a(file_path, sheet_name):
    # Load the workbook and select the sheet
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Iterate through each row in the sheet
    max_row = sheet.max_row
    for row in range(1, max_row + 1):
        # Get the value of column A
        value_a = sheet[f"A{row}"].value or ""

        # Concatenate the values from columns B, C, and D with line breaks
        value_b = sheet[f"B{row}"].value or ""
        value_c = sheet[f"C{row}"].value or ""
        # value_d = sheet[f"D{row}"].value or ""

        # Append the values of B, C, and D to A with line breaks
        concatenated_value = f"{value_a}\n{value_b}\n{value_c}"
        sheet[f"A{row}"] = concatenated_value  # Write back to column A

    # Save the updated workbook
    workbook.save(file_path)
    print(
        f"Data from columns B, C, and D has been appended to column A in sheet '{sheet_name}'."
    )


# Example usage
file_path = "BC_to_ab.xlsx"  # Replace with the path to your Excel file
sheet_name = "Sheet1"  # Replace with the sheet name

append_columns_bcd_to_a(file_path, sheet_name)
